 # pip install openpyxl
 # 참고 사이트 https://openpyxl.readthedocs.io/en/stable/tutorial.html


from openpyxl import Workbook, load_workbook

wb = Workbook()
ws = wb.active
ws.title = 'sample'
ws['B2'] = 42
ws.append([1, 2, 3])
ws.append([1, 2, 3])
ws.append([1, 2, 3])
ws.append([4, 4, 4])
wb.save('openpyxl1.xlsx')
wb.close()

wb = load_workbook(filename='openpyxl1.xlsx')
ws = wb.active
ws['A1'] = 42
ws.cell(row=1, column=3).value = 333
ws.append(['aaa', 'bbb', 'ccc'])

print(ws['A1'].value)
print(ws['A2'].value)

ws2 = wb['sample']
print(ws2['A3'].value, ws2['B3'].value, ws2['C3'].value)
print(ws2['A4'].value, ws2['B4'].value, ws2['C4'].value)
print(ws2['A5'].value, ws2['B5'].value, ws2['C5'].value)

wb.save('openpyxl1.xlsx')
wb.close()

# ==============
from openpyxl import Workbook

wb = Workbook()
# ws = wb.active
ws = wb.create_sheet("diary",0)

data = [('홍길동', 80,70,90), ('이기자', 90, 60 ,80), ('강기자', 80,80,70)]
r = 1
c = 1
for irum, kor, eng, math in data:
    ws.cell(row=r, column=c).value = irum
    ws.cell(row=r, column=c+1).value = kor
    ws.cell(row=r, column=c+2).value = eng
    ws.cell(row=r, column=c+3).value = math
    r += 1

ws.cell(row=r, column=1).value = '합계'
ws.cell(row=r, column=2).value = '=sum(B1:B3)'
ws.cell(row=r, column=3).value = '=sum(C1:C3)'
ws.cell(row=r, column=4).value = '=sum(D1:D3)'

wb.save('openpyxl2.xlsx')
wb.close()

#========
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.chart import BarChart, Reference

wb = Workbook()
ws = wb.create_sheet('chart', 0)
ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=4)
ws['A1'] = "성적표"
ws['A1'].font = Font(name='맑은 고딕', size=15, bold=True)
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

ws.append(['이름', '국어', '영어', '수학'])
ws.append(['홍길동', 60,70,60])
ws.append(['이기자', 88,77,89])
ws.append(['김기자', 55,66,77])

wb.save('openpyxl1_chart.xlsx')

barchart = BarChart()
barchart.title = "성적표"
barchart.x_axis.title = "이름"
barchart.y_axis.title = "점수"

data = Reference(ws, min_col=2, max_col=4, min_row=2, max_row=5)
cate = Reference(ws, min_col=1, max_col=1, min_row=3, max_row=5)

barchart.add_data(data, titles_from_data=True)
barchart.set_categories(cate)
barchart.style = 2

ws.add_chart(barchart, 'F1')
wb.save('openpyxl1_chart.xlsx')
wb.close()

